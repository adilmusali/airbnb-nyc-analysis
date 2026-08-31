from __future__ import annotations

import argparse
import json
import re
import urllib.request
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

DATA_INDEX_URL = "http://insideairbnb.com/get-the-data.html"
CITY_SLUG = "united-states/ny/new-york-city"
LISTINGS_REL = "visualisations/listings.csv"
DATA_HOST = "https://data.insideairbnb.com"

DATA_DIR = Path(__file__).resolve().parent / "data"
META_PATH = DATA_DIR / "snapshot_meta.json"
RNG = np.random.default_rng(42)
USER_AGENT = "airbnb-nyc-analysis/1.0 (+local data refresh)"


def _http_get(url: str, timeout: int = 60) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def find_latest_snapshot_date() -> str:
    """Parse Inside Airbnb's data index for the newest NYC snapshot date."""
    print(f"Checking for latest NYC snapshot via:\n  {DATA_INDEX_URL}")
    html = _http_get(DATA_INDEX_URL).decode("utf-8", errors="replace")
    dates = set(re.findall(r"new-york-city/(\d{4}-\d{2}-\d{2})/", html))
    if not dates:
        raise RuntimeError(
            "Could not find any New York City snapshot dates on the Inside Airbnb "
            "data page. Pass --snapshot YYYY-MM-DD to pin a known date."
        )
    latest = max(dates)  # ISO dates sort chronologically as strings
    print(f"Latest snapshot found: {latest}")
    return latest


def listings_url_for(snapshot: str) -> str:
    return f"{DATA_HOST}/{CITY_SLUG}/{snapshot}/{LISTINGS_REL}"


def load_snapshot_meta() -> dict | None:
    if not META_PATH.exists():
        return None
    try:
        return json.loads(META_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def save_snapshot_meta(snapshot: str, url: str, rows: int) -> None:
    META_PATH.write_text(
        json.dumps(
            {
                "city": "new-york-city",
                "snapshot": snapshot,
                "url": url,
                "rows": rows,
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def raw_files_exist() -> bool:
    required = (
        DATA_DIR / "airbnb_price.csv",
        DATA_DIR / "airbnb_room_type.xlsx",
        DATA_DIR / "airbnb_last_review.tsv",
    )
    return all(path.exists() for path in required)


def download_listings(url: str) -> pd.DataFrame:
    print(f"Downloading listings from:\n  {url}")
    df = pd.read_csv(url)
    print(f"Downloaded {len(df):,} rows, {len(df.columns)} columns")
    return df


def mess_up_room_type(series: pd.Series) -> pd.Series:
    """Apply inconsistent casing so the notebook has real cleaning work."""
    styles = ["title", "upper", "lower", "as_is"]
    choices = RNG.choice(styles, size=len(series))
    out = []
    for value, style in zip(series, choices):
        if pd.isna(value):
            out.append(value)
            continue
        text = str(value)
        if style == "title":
            out.append(text.title())
        elif style == "upper":
            out.append(text.upper())
        elif style == "lower":
            out.append(text.lower())
        else:
            out.append(text)
    return pd.Series(out, index=series.index, name=series.name)


def format_review_date(value) -> str | float:
    if pd.isna(value):
        return np.nan
    dt = pd.to_datetime(value, errors="coerce")
    if pd.isna(dt):
        return np.nan
    # e.g. "August 3 2026" (no comma, day without leading zero)
    return f"{dt.strftime('%B')} {dt.day} {dt.year}"


def listing_ids_as_text(df: pd.DataFrame) -> pd.Series:
    """Keep IDs as strings so Excel does not round 18+ digit values."""
    return df["id"].astype("int64").astype(str)


def build_price_csv(df: pd.DataFrame) -> pd.DataFrame:
    price = df["price"].copy()
    # Keep missing prices as missing; otherwise append " dollars"
    price_text = price.apply(
        lambda x: f"{int(x)} dollars" if pd.notna(x) else np.nan
    )
    nbhood_full = (
        df["neighbourhood_group"].astype(str)
        + ", "
        + df["neighbourhood"].astype(str)
    )
    # If either part was NaN, pandas turns it into the string "nan" — restore missing
    missing_nbhood = df["neighbourhood_group"].isna() | df["neighbourhood"].isna()
    nbhood_full = nbhood_full.mask(missing_nbhood, np.nan)

    return pd.DataFrame(
        {
            "listing_id": listing_ids_as_text(df),
            "price": price_text,
            "nbhood_full": nbhood_full,
        }
    )


def build_room_type_xlsx(df: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "listing_id": listing_ids_as_text(df),
            "description": df["name"],
            "room_type": mess_up_room_type(df["room_type"]),
        }
    )


def build_last_review_tsv(df: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "listing_id": listing_ids_as_text(df),
            "host_name": df["host_name"],
            "last_review": df["last_review"].map(format_review_date),
        }
    )


def write_excel_with_text_ids(room_df: pd.DataFrame, room_path: Path) -> None:
    room_df.to_excel(room_path, index=False)
    # Force listing_id to Excel text so large IDs are not rounded
    wb = load_workbook(room_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    id_col = headers.index("listing_id") + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=id_col)
        cell.number_format = "@"
        cell.value = str(cell.value)
    wb.save(room_path)


def prepare(snapshot: str | None = None, force: bool = False) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    snapshot = snapshot or find_latest_snapshot_date()
    url = listings_url_for(snapshot)
    meta = load_snapshot_meta()

    if (
        not force
        and meta
        and meta.get("snapshot") == snapshot
        and raw_files_exist()
    ):
        print(
            f"Already up to date with snapshot {snapshot}. "
            "Use --force to re-download and rebuild."
        )
        return

    listings = download_listings(url)

    price_df = build_price_csv(listings)
    room_df = build_room_type_xlsx(listings)
    review_df = build_last_review_tsv(listings)

    price_path = DATA_DIR / "airbnb_price.csv"
    room_path = DATA_DIR / "airbnb_room_type.xlsx"
    review_path = DATA_DIR / "airbnb_last_review.tsv"

    price_df.to_csv(price_path, index=False)
    write_excel_with_text_ids(room_df, room_path)
    review_df.to_csv(review_path, sep="\t", index=False)
    save_snapshot_meta(snapshot, url, len(listings))

    print(f"Wrote {price_path}  ({len(price_df):,} rows)")
    print(f"Wrote {room_path}  ({len(room_df):,} rows)")
    print(f"Wrote {review_path}  ({len(review_df):,} rows)")
    print(f"Recorded snapshot metadata in {META_PATH}")
    print("\nSample price values:", price_df["price"].dropna().head(3).tolist())
    print("Sample room_type values:", room_df["room_type"].dropna().unique()[:6].tolist())
    print("Sample last_review values:", review_df["last_review"].dropna().head(3).tolist())
    print("Sample nbhood_full:", price_df["nbhood_full"].dropna().head(2).tolist())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download the latest (or pinned) Inside Airbnb NYC snapshot "
        "and rebuild the three raw analysis files."
    )
    parser.add_argument(
        "--snapshot",
        metavar="YYYY-MM-DD",
        help="Pin a specific snapshot date instead of auto-detecting the latest.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Rebuild even when local files already match the target snapshot.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    prepare(snapshot=args.snapshot, force=args.force)


if __name__ == "__main__":
    main()
